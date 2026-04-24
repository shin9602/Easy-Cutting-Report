"""
절삭평가 이미지 세트 재조립 도구 v2.5 (Visual Integrity Edition)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 엑셀의 이미지 Span(Colspan/Rowspan)을 정확히 파싱하여 반영
- 0-갭(Zero-Gap) 밀착 그리드 레이아웃
- 이미지 100% 가득 채우기 (여백 근절)
- R부/크레이터 제트인덱스(z-index) 레이어링
- 상세 로깅 시스템 (debug_layout.log)
"""

import sys, os, io, zipfile, base64, re, argparse
import xml.etree.ElementTree as ET
from collections import defaultdict
from PIL import Image
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import webbrowser, tempfile
import subprocess

# ──── 상수 ────
NS_WS   = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
NS_A    = 'http://schemas.openxmlformats.org/drawingml/2006/main'
NS_R    = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
R_EMBED = '{' + NS_R + '}embed'
EMU_PX  = 9525
PT_PX   = 4 / 3


# ═══════════════════════════════════════════
# 전처리 및 파싱
# ═══════════════════════════════════════════

_XLS_CONVERTER_SCRIPT = r"""
import sys, os, time
abs_xls, abs_out = sys.argv[1], sys.argv[2]

# COM initialization for Excel
import pythoncom, win32com.client
try:
    pythoncom.CoInitialize()
except:
    pass

# Create Excel application with proper error handling
excel = None
try:
    excel = win32com.client.Dispatch("Excel.Application")
    if excel is None:
        raise RuntimeError("Excel.Application 초기화 실패")
    
    # Set properties
    try:
        excel.Visible = False
    except:
        pass
    try:
        excel.DisplayAlerts = False
    except:
        pass
    
    # Ensure Workbooks attribute exists
    wb = None
    try:
        wb = excel.Workbooks.Open(abs_xls)
        wb.SaveAs(abs_out, FileFormat=51)
        wb.Close(SaveChanges=False)
    finally:
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
    
    excel.Quit()
except AttributeError as e:
    if "Workbooks" in str(e):
        raise RuntimeError("Excel COM 초기화 실패: Workbooks 속성에 접근할 수 없습니다. Excel이 제대로 설치되지 않았을 수 있습니다.")
    raise
finally:
    try:
        if excel:
            excel.Quit()
    except:
        pass
    try:
        pythoncom.CoUninitialize()
    except:
        pass
"""

def _xls_to_xlsx_python_fallback(xls_path, out_path):
    """Pure Python fallback: xlrd + openpyxl를 사용하여 .xls를 .xlsx로 변환"""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Read XLS file
    xls_wb = xlrd.open_workbook(xls_path, formatting_info=True)
    
    # Create new XLSX workbook
    xlsx_wb = Workbook()
    xlsx_wb.remove(xlsx_wb.active)
    
    for sheet_idx in range(xls_wb.nsheets):
        xls_sheet = xls_wb.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_wb.create_sheet(title=xls_sheet.name if xls_sheet.name else f"Sheet{sheet_idx}")
        
        # Copy cell values and basic formatting
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                if cell_value:
                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
    
    # Save XLSX file
    xlsx_wb.save(out_path)

def xls_to_xlsx(xls_path):
    out_path = os.path.splitext(xls_path)[0] + '.xlsx'
    if os.path.exists(out_path): return out_path
    abs_xls = os.path.abspath(xls_path)
    abs_out = os.path.abspath(out_path)

    # First, try using Excel COM via subprocess
    try:
        import subprocess, tempfile as _tempfile
        script_fd, script_path = _tempfile.mkstemp(suffix='.py')
        try:
            with os.fdopen(script_fd, 'w', encoding='utf-8') as f:
                f.write(_XLS_CONVERTER_SCRIPT)
            result = subprocess.run(
                [sys.executable, script_path, abs_xls, abs_out],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0 and os.path.exists(abs_out):
                return abs_out
            # COM 실패 시 fallback으로 진행 (에러 무시)
        finally:
            try: os.remove(script_path)
            except: pass
    except Exception:
        pass

    # Fallback: use pure Python xlrd + openpyxl
    try:
        _xls_to_xlsx_python_fallback(abs_xls, abs_out)
        if os.path.exists(abs_out):
            return abs_out
        raise RuntimeError("파일 생성 실패")
    except ImportError as e:
        raise RuntimeError(
            f".xls → .xlsx 변환 실패: {e}\n\n"
            "다음 패키지를 설치하세요:\n"
            "  pip install xlrd openpyxl"
        )
    except Exception as e:
        raise RuntimeError(
            f".xls → .xlsx 변환 실패: {e}\n\n"
            "1. Microsoft Excel이 설치되어 있는지 확인하세요\n"
            "2. 또는 xlrd/openpyxl 패키지를 설치하세요: pip install xlrd openpyxl"
        )


def xls_to_xlsx_old(xls_path):
    out_path = os.path.splitext(xls_path)[0] + '.xlsx'
    if os.path.exists(out_path): return out_path
    abs_xls = os.path.abspath(xls_path)
    abs_out = os.path.abspath(out_path)
    try:
        import subprocess, tempfile
        script_fd, script_path = tempfile.mkstemp(suffix='.py')
        try:
            with os.fdopen(script_fd, 'w', encoding='utf-8') as f:
                f.write(_XLS_CONVERTER_SCRIPT)
            result = subprocess.run(
                [sys.executable, script_path, abs_xls, abs_out],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode != 0:
                raise RuntimeError(result.stderr.strip() or result.stdout.strip())
        finally:
            try: os.remove(script_path)
            except: pass
        if not os.path.exists(abs_out):
            raise RuntimeError("변환 후 파일이 생성되지 않았습니다.")
        return abs_out
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f".xls → .xlsx 자동 변환 실패: {e}\n\nMicrosoft Excel이 설치되어 있는지 확인하세요.")


def parse_xls(xls_path):
    import xlrd
    wb = xlrd.open_workbook(xls_path, formatting_info=True)
    sh = wb.sheet_by_index(0)
    row_h = {r: int(info.height / 20 * PT_PX) for r, info in sh.rowinfo_map.items()}
    col_w = {c: int(info.width / 256 * 8) for c, info in sh.colinfo_map.items()}
    for r in range(sh.nrows):
        if r not in row_h: row_h[r] = 20
    for c in range(sh.ncols):
        if c not in col_w: col_w[c] = 64
    text_cells = {}
    for r in range(sh.nrows):
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if v: text_cells[(r, c)] = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
    return row_h, col_w, text_cells, sh.merged_cells, sh.nrows, sh.ncols


def parse_xls_from_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    n_rows, n_cols = ws.max_row, ws.max_column
    row_h, col_w = {}, {}
    for r in range(n_rows):
        rh = ws.row_dimensions.get(r + 1)
        row_h[r] = int(rh.height * PT_PX) if rh and rh.height else 20
    for c in range(n_cols):
        cw = ws.column_dimensions.get(chr(65 + c) if c < 26 else None)
        col_w[c] = int(cw.width * 8) if cw and cw.width else 64
    text_cells = {}
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            v = ws.cell(r, c).value
            if v is not None: text_cells[(r - 1, c - 1)] = str(v).strip()
    merges = []
    for mg in ws.merged_cells.ranges:
        merges.append((mg.min_row - 1, mg.max_row, mg.min_col - 1, mg.max_col))
    return row_h, col_w, text_cells, merges, n_rows, n_cols


def parse_images(xlsx_path):
    """
    이미지의 Span 정보를 파악하여 엑셀에서의 실제 크기를 유지하도록 함
    """
    img_data, raw = {}, []
    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()
        for name in names:
            if 'xl/media/' in name: img_data[name.split('/')[-1]] = z.read(name)
        drw_files = [n for n in names if 'xl/drawings/drawing' in n and n.endswith('.xml')]
        for drw_xml in drw_files:
            rels_path = drw_xml.replace('drawings/', 'drawings/_rels/') + '.rels'
            if rels_path not in names: continue
            rid_map = {rel.get('Id'): rel.get('Target', '').split('/')[-1] for rel in ET.fromstring(z.read(rels_path).decode())}
            root = ET.fromstring(z.read(drw_xml).decode())
            for anchor in root:
                tag = anchor.tag.split('}')[-1]
                if tag not in ('twoCellAnchor', 'oneCellAnchor'): continue
                
                frm = anchor.find('{' + NS_WS + '}from')
                to = anchor.find('{' + NS_WS + '}to')
                if frm is None: continue
                
                r1, c1 = int(frm.find('{' + NS_WS + '}row').text), int(frm.find('{' + NS_WS + '}col').text)
                rs, cs = 1, 1
                if to is not None:
                    r2, c2 = int(to.find('{' + NS_WS + '}row').text), int(to.find('{' + NS_WS + '}col').text)
                    rs, cs = max(1, r2 - r1 + 1), max(1, c2 - c1 + 1)
                
                pic = anchor.find('{' + NS_WS + '}pic')
                if pic is None: continue
                blip = pic.find('.//{' + NS_A + '}blip')
                if blip is None: continue
                
                cx, cy = 0, 0
                for child in anchor.iter():
                    if child.tag.split('}')[-1] == 'ext' and child.get('cx'):
                        cx, cy = int(child.get('cx')) // EMU_PX, int(child.get('cy')) // EMU_PX; break
                
                crop = {'l':0.0, 't':0.0, 'r':0.0, 'b':0.0}
                srcRect = pic.find('.//{' + NS_A + '}srcRect')
                if srcRect is not None:
                    for k in ('l','t','r','b'):
                        v = srcRect.get(k)
                        if v: crop[k] = float(v)/100000.0
                
                fname = rid_map.get(blip.get(R_EMBED), '')
                if fname: raw.append((r1, c1, rs, cs, cx, cy, fname, crop))

    cell_map = defaultdict(list)
    for r, c, rs, cs, cx, cy, fname, crop in raw: 
        cell_map[(r, c)].append({'fname': fname, 'rs': rs, 'cs': cs, 'cx': cx, 'cy': cy, 'crop': crop})
    
    img_cells = {}
    for (r, c), lst in cell_map.items():
        lst.sort(key=lambda x: x['cx']*x['cy'], reverse=True)
        if len(lst) == 1:
            inf = lst[0]
            img_cells[(r, c)] = [{'fname': inf['fname'], 'rs': inf['rs'], 'cs': inf['cs'], 'disp_w': inf['cx'], 'disp_h': inf['cy'], 'crop_pct': inf['crop']}]
        else:
            # 겹침 처리
            big, sml = lst[0], lst[1]
            img_cells[(r, c)] = [
                {'fname': sml['fname'], 'rs': sml['rs'], 'cs': sml['cs'], 'disp_w': sml['cx'], 'disp_h': sml['cy'], 'crop_pct': sml['crop'], 'is_r_part': True},
                {'fname': big['fname'], 'rs': big['rs'], 'cs': big['cs'], 'disp_w': big['cx'], 'disp_h': big['cy'], 'crop_pct': big['crop'], 'is_crater': True}
            ]
    return img_cells, img_data


def img_to_b64(raw_bytes, info, quality, cell_w, cell_h):
    crop = info.get('crop_pct', {'l':0.0, 't':0.0, 'r':0.0, 'b':0.0})
    im = Image.open(io.BytesIO(raw_bytes)).convert('RGB')
    ow, oh = im.size
    cl, ct = max(0, int(ow*crop['l'])), max(0, int(oh*crop['t']))
    cr, cb = min(ow, int(ow*(1.0-crop['r']))), min(oh, int(oh*(1.0-crop['b'])))
    if cl < cr and ct < cb: im = im.crop((cl, ct, cr, cb))
    
    tw = max(cell_w, 1)
    sw, sh = im.size
    ratio = tw / sw if sw > 0 else 1.0
    th = max(1, int(sh * ratio))
    im = im.resize((tw, th), Image.LANCZOS)
    
    with open("debug_layout.log", "a", encoding="utf-8") as f:
        f.write(f"[IMG] {info['fname']} | Orig:{ow}x{oh} | Crop({cl},{ct},{cr},{cb}) | Final:{tw}x{th}\n")

    buf = io.BytesIO()
    im.save(buf, 'JPEG', quality=quality, subsampling=0)
    return base64.b64encode(buf.getvalue()).decode(), th


# ═══════════════════════════════════════════
# 세트 분석
# ═══════════════════════════════════════════

def analyze_sets(text_cells, img_cells, n_rows, n_cols):
    pass_info = {}
    for (r, c), txt in text_cells.items():
        if c == 0:
            m = re.match(r'(\d+)\s*P', txt, re.IGNORECASE)
            if m: pass_info[r] = int(m.group(1))
            
    insert_cols = {}
    for (r, c), txt in text_cells.items():
        # 홀수 열(1, 3, 5, 7, 9, ...)을 모두 데이터 시작 열로 간주 (무제한)
        if c % 2 == 1:
            if c not in insert_cols: insert_cols[c] = {}
            if r == 1: insert_cols[c]['num'] = txt
            elif r == 2: insert_cols[c]['chip'] = txt
            elif r == 3: insert_cols[c]['grade'] = txt
            
    pass_rows = sorted(pass_info.keys())
    # ... 나머지 로직은 그대로
    pass_ranges = {pass_info[pr]: (pr, pass_rows[i+1] if i+1 < len(pass_rows) else n_rows) for i, pr in enumerate(pass_rows)}
    
    sets = {}
    for pn, (rs, re_idx) in pass_ranges.items():
        for cg in sorted(insert_cols.keys()):
            cs, ce = cg, cg + 2
            has_img = any((r,c) in img_cells for r in range(rs, re_idx) for c in range(cs, ce))
            if has_img:
                info = insert_cols.get(cg, {})
                sets[f"{pn}_{cg}"] = {
                    'pass': pn, 'grade': info.get('grade',''), 'chip': info.get('chip',''), 'num': info.get('num',''),
                    'row_range': (rs, re_idx), 'col_range': (cs, ce)
                }
    return sets


def _build_local_span_map(merges, rs, re, cs, ce, img_cells):
    span_map, skip = {}, set()
    # 엑셀 병합 셀 반영
    for rlo, rhi, clo, chi in merges:
        if rs <= rlo < re and cs <= clo < ce:
            ars, acs = min(rhi-rlo, re-rlo), min(chi-clo, ce-clo)
            span_map[(rlo, clo)] = (ars, acs)
            for rr in range(rlo, rlo+ars):
                for cc in range(clo, clo+acs):
                    if (rr, cc) != (rlo, clo): skip.add((rr, cc))
    # 이미지 자체의 Span 반영 (병합 셀이 아닌 경우를 대비)
    for (ir, ic), ilist in img_cells.items():
        if rs <= ir < re and cs <= ic < ce:
            inf = ilist[0]
            ars, acs = min(inf['rs'], re-ir), min(inf['cs'], ce-ic)
            if ars > 1 or acs > 1:
                cur_s = span_map.get((ir, ic), (1, 1))
                span_map[(ir, ic)] = (max(cur_s[0], ars), max(cur_s[1], acs))
                for rr in range(ir, ir+ars):
                    for cc in range(ic, ic+acs):
                        if (rr, cc) != (ir, ic): skip.add((rr, cc))
    return span_map, skip


# ═══════════════════════════════════════════
# HTML 엔진 (Visual Integrity Version)
# ═══════════════════════════════════════════

def build_subset_html(selected_sets, row_h, col_w, text_cells, merges, img_cells, img_data, quality, scale, title):
    if not selected_sets: return "<html><body>No Data Selected</body></html>"
    
    with open("debug_layout.log", "w", encoding="utf-8") as f:
        f.write("--- HTML REASSEMBLY DEBUG ---\n")

    p_scale = scale * 1.0
    col_groups = defaultdict(list)
    for s in selected_sets: col_groups[s['col_range'][0]].append(s)
    
    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 20px; background: #eef2f5; font-family: 'Malgun Gothic', sans-serif; }}
  .toolbar {{ position: sticky; top: 0; background: #1a1a2e; color: #fff; padding: 12px 20px; z-index: 10000; display: flex; gap: 15px; align-items: center; border-radius: 0 0 8px 8px; margin-bottom: 20px; }}
  .btn {{ background: #00b894; color: #fff; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-weight: bold; transition: 0.2s; }}
  .btn:hover {{ background: #009376; transform: translateY(-1px); }}
  #table-wrap {{ background: #fff; padding: 15px; margin: 0 auto; width: fit-content; border-radius: 4px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }}
  .tetris-grid {{ display: grid; gap: 0; width: 100%; }}
  .column-stack {{ display: flex; flex-direction: column; gap: 0; border: 1px solid #000; margin-right: -1px; }}
  .grade-header {{ background: #2d3436; color: #fff; text-align: center; font-size: 16px; font-weight: bold; padding: 12px; border-bottom: 2px solid #000; text-transform: uppercase; }}
  .set-box {{ margin: 0; border-bottom: 1px solid #000; }}
  table {{ border-collapse: collapse; width: 100%; table-layout: fixed; margin: 0; padding: 0; border: none; }}
  td {{ border: 1px solid #000; padding: 0; margin: 0; vertical-align: top; overflow: visible; }}
  .img-container {{ line-height: 0; position: relative; width: 100%; }}
  .img-item {{ width: 100%; height: auto; display: block; }}
  .pass-label {{ background: #000; color: #fff; text-align: center; font-weight: bold; padding: 10px; font-size: 17px; }}
  @media print {{ .toolbar {{ display: none; }} body {{ background: #fff; padding: 0; }} #table-wrap {{ box-shadow: none; padding: 0; }} }}
</style>
</head>
<body>
<div class="toolbar">
  <span style="font-size: 18px; font-weight: bold;">{title}</span>
  <button class="btn" onclick="captureAll()">📷 고화질 이미지 추출</button>
  <button class="btn" onclick="window.print()" style="background:#0984e3;">📄 PDF 리포트 생성</button>
</div>
<div id="table-wrap">
  <div class="tetris-grid" style="grid-template-columns: repeat({len(col_groups)}, 1fr);">
"""
    
    for cs_idx in sorted(col_groups.keys()):
        column_sets = sorted(col_groups[cs_idx], key=lambda x: x['pass'])
        grade_name = column_sets[0]['grade']
        html += f'<div class="column-stack"><div class="grade-header">{grade_name}</div>'
        
        for s in column_sets:
            rs, re, cs, ce = s['row_range'][0], s['row_range'][1], s['col_range'][0], s['col_range'][1]
            pass_str = f"{s['pass']}P"
            
            # 여기서 이미지의 Span을 포함하는 로직 사용
            span_map, skip = _build_local_span_map(merges, rs, re, cs, ce, img_cells)
            col_tags = ''.join(f'<col style="width:{int(col_w.get(c, 64)*p_scale)}px">' for c in range(cs, ce))
            
            # 가장 사진이 적은 열(측면도) 아래에 라벨 배치
            side_c = -1
            col_imgs = defaultdict(int)
            for r_s in range(rs, re):
                for c_s in range(cs, ce):
                    if (r_s, c_s) in img_cells: col_imgs[c_s] += len(img_cells[(r_s, c_s)])
            if col_imgs: side_c = min(col_imgs, key=col_imgs.get)
            
            rows_html, pass_done = "", False
            for r in range(rs, re):
                has_img_row = any((r,c) in img_cells for c in range(cs, ce))
                style_tr = "" if has_img_row else f' style="height:{int(row_h.get(r, 20)*p_scale)}px"'
                rows_html += f'<tr{style_tr}>'
                for c in range(cs, ce):
                    if (r,c) in skip: continue
                    rspan, cspan = span_map.get((r,c), (1,1))
                    attrs = (f' rowspan="{rspan}"' if rspan>1 else '') + (f' colspan="{cspan}"' if cspan>1 else '')
                    bw = sum(int(col_w.get(c+i, 64)*p_scale) for i in range(cspan))
                    
                    cell_content, t_style = "", f"width:{bw}px; background:#fff;"
                    if (r,c) in img_cells:
                        imgs = img_cells[(r,c)]
                        crater = next((i for i in imgs if i.get('is_crater')), None)
                        r_part = next((i for i in imgs if i.get('is_r_part')), None)
                        
                        if crater and r_part:
                            # 겹침 모드 (Absolute Layering)
                            c64, ch = img_to_b64(img_data[crater['fname']], crater, quality, bw, 20)
                            rw_ratio = r_part['disp_w']/crater['disp_w']
                            r64, rh = img_to_b64(img_data[r_part['fname']], r_part, quality, int(bw*rw_ratio), 20)
                            cell_content = f"""<div class="img-container" style="height:{ch}px; overflow:hidden;">
                                <img src="data:image/jpeg;base64,{c64}" class="img-item">
                                <img src="data:image/jpeg;base64,{r64}" style="position:absolute; top:0; right:0; width:{round(rw_ratio*100,1)}%; box-shadow: 0 0 5px rgba(0,0,0,0.5);">
                            </div>"""
                        else:
                            inner = ""
                            for inf in imgs:
                                b64, ih = img_to_b64(img_data[inf['fname']], inf, quality, bw, 20)
                                inner += f'<img src="data:image/jpeg;base64,{b64}" class="img-item">'
                            cell_content = f'<div class="img-container">{inner}</div>'
                        
                        if not pass_done and c == side_c:
                            cell_content += f'<div class="pass-label">{pass_str}</div>'
                            pass_done = True
                    elif (r,c) in text_cells:
                        cell_content = text_cells[(r,c)]
                        t_style += "background:#f1f2f6; text-align:center; font-weight:bold; font-size:12px; vertical-align:middle; padding:3px;"
                    
                    rows_html += f'<td{attrs} style="{t_style}">{cell_content}</td>'
                rows_html += "</tr>"
            
            html += f'<div class="set-box"><table>{col_tags}<tbody>{rows_html}</tbody></table></div>'
        html += '</div>'
        
    html += """</div></div>
<script>
function captureAll() {
  const el = document.getElementById('table-wrap');
  html2canvas(el, { scale: 3, useCORS: true }).then(c => {
    const a = document.createElement('a'); a.download = 'cutting_eval_packed_' + Date.now() + '.png'; a.href = c.toDataURL(); a.click();
  });
}
</script>
</body></html>"""
    return html


# ═══════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════

class CuttingEvalApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("절삭평가 재조립 마스터 v2.5")
        self.root.geometry("900x700")
        self.root.configure(bg='#2d3436')
        self.image_sets, self.img_cells, self.img_data, self.text_cells = {}, {}, {}, {}
        self.row_h, self.col_w, self.merges, self.n_rows, self.n_cols = {}, {}, [], 0, 0
        self.check_vars = {}
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg='#2d3436'); hdr.pack(fill='x', padx=20, pady=20)
        tk.Label(hdr, text="💎 절삭평가 자동화 리포트 도구", bg='#2d3436', fg='#00b894', font=('맑은 고딕', 18, 'bold')).pack(side='left')
        
        mf = tk.Frame(self.root, bg='#636e72', padx=15, pady=15); mf.pack(fill='x', padx=20)
        self.file_e = tk.Entry(mf, width=55, font=('맑은 고딕', 10)); self.file_e.pack(side='left', padx=5)
        tk.Button(mf, text="파일 선택", command=self._sel_f, bg='#dfe6e9', relief='flat').pack(side='left', padx=5)
        tk.Button(mf, text="데이터 불러오기", command=self._load, bg='#00b894', fg='#fff', relief='flat', padx=10, font=('맑은 고딕',10,'bold')).pack(side='left', padx=5)
        
        self.list_canvas = tk.Canvas(self.root, bg='#fff', height=350)
        self.vsb = ttk.Scrollbar(self.root, orient="vertical", command=self.list_canvas.yview)
        self.scroll_frame = tk.Frame(self.list_canvas, bg='#fff')
        self.scroll_frame.bind("<Configure>", lambda e: self.list_canvas.configure(scrollregion=self.list_canvas.bbox("all")))
        self.list_canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.list_canvas.configure(yscrollcommand=self.vsb.set)
        self.list_canvas.pack(side="left", fill="both", expand=True, padx=(20,0), pady=10)
        self.vsb.pack(side="right", fill="y", padx=(0,20), pady=10)
        
        self.title_e = tk.Entry(self.root, font=('맑은 고딕', 12)); self.title_e.pack(fill='x', padx=20, pady=10)
        self.title_e.insert(0, "절삭평가 결과 보고서")
        
        btn_f = tk.Frame(self.root, bg='#2d3436'); btn_f.pack(pady=20)
        tk.Button(btn_f, text="🔍 시각적 레이아웃 미리보기", command=self._html, bg='#e17055', fg='#fff', font=('맑은 고딕', 12, 'bold'), padx=20, pady=10, relief='flat').pack()

    def _sel_f(self):
        p = filedialog.askopenfilename(); self.file_e.delete(0, tk.END); self.file_e.insert(0, p)

    def _kill_excel_com_users(self):
        """Excel COM을 사용 중인 다른 프로세스를 찾아 종료 여부를 묻고 처리"""
        try:
            import wmi
            c = wmi.WMI()
            excel_procs = [p for p in c.Win32_Process() if p.Name and p.Name.upper() == 'EXCEL.EXE']
        except Exception:
            # wmi 없으면 tasklist raw bytes로 시도
            try:
                result = subprocess.run(
                    ['tasklist', '/FI', 'IMAGENAME eq EXCEL.EXE', '/NH'],
                    capture_output=True, timeout=10
                )
                out = result.stdout.decode('cp949', errors='ignore')
                excel_procs = None  # wmi 객체 대신 pid 목록으로 처리
                pids = []
                for line in out.splitlines():
                    parts = line.split()
                    if parts and parts[0].upper() == 'EXCEL.EXE' and len(parts) >= 2:
                        pids.append(parts[1])
                if not pids:
                    return True
                answer = messagebox.askyesno(
                    "Excel 충돌 감지",
                    f"Excel이 실행 중입니다 (PID: {', '.join(pids)}).\n\n"
                    "Excel을 종료하고 계속 진행할까요?\n"
                    "(저장 안 된 작업이 있으면 손실될 수 있습니다)"
                )
                if not answer:
                    return False
                for pid in pids:
                    subprocess.run(['taskkill', '/PID', pid, '/F'], capture_output=True)
                import time; time.sleep(1.5)
                return True
            except Exception:
                return True

        if not excel_procs:
            return True

        pid_list = ', '.join([str(p.ProcessId) for p in excel_procs])
        answer = messagebox.askyesno(
            "Excel 충돌 감지",
            f"Excel이 실행 중입니다 (PID: {pid_list}).\n\n"
            "Excel을 종료하고 계속 진행할까요?\n"
            "(저장 안 된 작업이 있으면 손실될 수 있습니다)"
        )
        if not answer:
            return False
        for p in excel_procs:
            subprocess.run(['taskkill', '/PID', str(p.ProcessId), '/F'], capture_output=True)
        import time; time.sleep(1.5)
        return True

    def _load(self):
        p = self.file_e.get().strip()
        if not p: return
        try:
            ext = os.path.splitext(p)[1].lower()
            if ext == '.xls':
                if not self._kill_excel_com_users():
                    return
                self.row_h, self.col_w, self.text_cells, self.merges, self.n_rows, self.n_cols = parse_xls(p)
                xp = xls_to_xlsx(p)
            else:
                xp = p
                self.row_h, self.col_w, self.text_cells, self.merges, self.n_rows, self.n_cols = parse_xls_from_xlsx(p)
            self.img_cells, self.img_data = parse_images(xp)
            self.image_sets = analyze_sets(self.text_cells, self.img_cells, self.n_rows, self.n_cols)
            for w in self.scroll_frame.winfo_children(): w.destroy()
            self.check_vars = {}
            for k in sorted(self.image_sets.keys()):
                s = self.image_sets[k]
                v = tk.BooleanVar(value=True); self.check_vars[k] = v
                tk.Checkbutton(self.scroll_frame, text=f"  {s['grade']} | {s['pass']}P | {s['num']}", variable=v, bg='#fff').pack(anchor='w', padx=10, pady=2)
        except Exception as e: messagebox.showerror("Load Failed", str(e))

    def _html(self):
        sel = [self.image_sets[k] for k, v in self.check_vars.items() if v.get()]
        if not sel: return
        h = build_subset_html(sel, self.row_h, self.col_w, self.text_cells, self.merges, self.img_cells, self.img_data, 95, 1.3, self.title_e.get())
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.html', mode='w', encoding='utf-8')
        tmp.write(h); tmp.close(); webbrowser.open(f'file:///{tmp.name}')

    def run(self): self.root.mainloop()

def get_all_data_json(xlsx_path):
    import json
    row_h, col_w, text_cells, merges, n_rows, n_cols = parse_xls_from_xlsx(xlsx_path)
    img_cells, img_data = parse_images(xlsx_path)
    image_sets = analyze_sets(text_cells, img_cells, n_rows, n_cols)
    
    # 이미지 데이터를 Base64로 미리 변환 (JSON용)
    b64_cache = {}
    for fname, data in img_data.items():
        b64_cache[fname] = base64.b64encode(data).decode()
        
    res = {
        'row_h': row_h, 'col_w': col_w, 'text_cells': {f"{k[0]},{k[1]}": v for k, v in text_cells.items()},
        'img_cells': {f"{k[0]},{k[1]}": v for k, v in img_cells.items()},
        'img_data_b64': b64_cache,
        'merges': merges,
        'image_sets': list(image_sets.values()),
        'n_rows': n_rows, 'n_cols': n_cols
    }
    return json.dumps(res, ensure_ascii=False)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--json', help='Excel file path to get JSON data')
    args = parser.parse_args()
    
    if args.json:
        # CLI 모드: JSON 출력
        try:
            print(get_all_data_json(args.json))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    else:
        # GUI 모드
        CuttingEvalApp().run()
